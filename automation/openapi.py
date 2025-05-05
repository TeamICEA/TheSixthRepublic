import requests
import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import time
import json

keys = {}
with open("keys.json") as f:
    keys = json.load(f)
conn=mysql.connector.connect(
host=keys["SQL_HOST"],
port=3306,
user=keys["SQL_USERNAME"],
password=keys["SQL_PASSWORD"],
database=keys["SQL_DATABASE"]
)
cursor=conn.cursor()

def parse_birthdate(birth_str):
    try:
        return datetime.strptime(birth_str,"%Y-%m-%d").date()
    except:
        return None

    
def get_partyid(party_name):
    if '더불어민주당' in party_name:
        return 1
    elif '국민의힘' in party_name:
        return 2
    elif '조국혁신당' in party_name:
        return 3
    elif '개혁신당' in party_name:
        return 4
    elif '진보당' in party_name:
        return 5
    elif '기본소득당' in party_name:
        return 6
    elif '사회민주당' in party_name:
        return 7
    else:
        return 0



API_KEY=keys["OPENAPI_KEY"]
openapi_url = f"https://open.assembly.go.kr/portal/openapi/ALLNAMEMBER?KEY={API_KEY}&Type=json"

page_size=1000

for page in range(1,5):
    res=requests.get(f"{openapi_url}&pIndex={page}&pSize={page_size}")
    data=res.json()
    items=data['ALLNAMEMBER'][1]['row']

    for p in items:
        if "제22대" not in (p.get('GTELT_ERACO') or ''):
            continue
        
        party_name=p.get('PLPT_NM','')
        party_id=get_partyid(party_name)

        if party_id==0:
            continue

        cursor.execute("""
        INSERT INTO politicians (
            party_id,name,hanja_name,english_name,birthdate_type,birthdate,job,
            party,gender,reelected,tel,email,profile,address,boja,top_secretary,
            secretatry,pic_link,comittees,str_id
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s)
    """, (
        party_id,
        p.get('NAAS_NM'),
        p.get('NAAS_CH_NM'),
        p.get('NAAS_EN_NM'),
        p.get('BIRDY_DIV_CD'),
        parse_birthdate(p.get('BIRDY_DT')),
        p.get('DTY_NM'),
        party_name,
        p.get('NTR_DIV'),
        p.get('RLCT_DIV_NM'),
        p.get('NAAS_TEL_NO'),
        p.get('NAAS_EMAIL_ADDR'),
        p.get('BRF_HST'),
        p.get('OFFM_RNUM_NO'),
        p.get('AIDE_NM'),
        p.get('CHF_SCRT_NM'),
        p.get('SCRT_NM'),
        p.get('NAAS_PIC'),
        p.get('BLNG_CMIT_NM'),
        p.get('NAAS_CD')
    ))
        
    conn.commit()



#오픈와치 api
url=f"https://www.openwatch.kr/api/national-assembly/members?age=22&pageSize=100"
for page in range(1,4):
    res=requests.get(f"{url}&page={page}")
    data=res.json()
    items=data['rows']

    for p in items:

        party_name=p.get('partyName','')
        party_id=get_partyid(party_name)
        mona_code=p.get('monaCode')
        if party_id==0:
            continue

        
        cursor.execute("SELECT COUNT(*) FROM politicians WHERE str_id=%s",(mona_code,))
        exists=cursor.fetchone()[0]
        if exists>0:
            cursor.execute("UPDATE politicians SET books = %s WHERE str_id = %s", (p.get('books'), mona_code))
        else:
            cursor.execute("""
            INSERT INTO politicians (
                party_id,name,hanja_name,english_name,birthdate_type,birthdate,job,
                party,gender,reelected,tel,email,profile,address,boja,top_secretary,
                secretatry,comittees,str_id,books
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s)
            """, (
            party_id,
            p.get('name'),
            p.get('nameHanja'),
            p.get('nameEnglish'),
            p.get('birthdateType'),
            parse_birthdate(p.get('birthdate')),
            p.get('job'),
            party_name,
            p.get('gender'),
            p.get('reelected'),
            p.get('tel'),
            p.get('email'),
            p.get('profile'),
            p.get('officeAddress'),
            p.get('staff'),
            p.get('secretary'),
            p.get('secretary2'),
            p.get('committees'),
            p.get('monaCode'),
            p.get('books')
        ))
        conn.commit()
            


#재산엑셀파일
wb = load_workbook(r"2023.03 국회고위공직자 정기재산신고.xlsx", data_only=True)
ws = wb.active  # 또는 wb["시트이름"]

# ─── 열 인덱스 자동 찾기 (1행을 기준으로) ───────
header = {cell.value: idx for idx, cell in enumerate(ws[1])}
mona_idx = header['monaCode']
curr_idx = header['현재가액']

# ─── 각 행을 순회하며 DB 업데이트 ───────────────
for row in ws.iter_rows(min_row=2, values_only=True):
    mona_code = str(row[mona_idx]).strip()
    curr_asset = row[curr_idx]

    if mona_code and curr_asset is not None:
        try:
            # 쉼표 제거하고 숫자형으로 변환
            asset_value = int(str(curr_asset).replace(',', '').strip())
            cursor.execute("UPDATE politicians SET curr_assets = %s WHERE str_id = %s", (asset_value, mona_code))
        except Exception as e:
            print(f"변환 실패 → monaCode: {mona_code}, curr_asset: {curr_asset}, error: {e}")

    conn.commit()




#찬성 법안
# 1. 기존 데이터 미리 불러오기
cursor.execute("SELECT str_id, bill_approved FROM politicians")
existing_data = {row[0]: row[1] or "" for row in cursor.fetchall()}
updates = {}

# 2. 찬성 법안 API 순회
url_bill = "https://www.openwatch.kr/api/national-assembly/votes?age=22&pageSize=100"

for page in range(1, 870):
    try:
        print(f"[{page}/869] 요청 중...")

        # 요청 시 재시도 로직 추가
        for retry in range(3):
            try:
                res = requests.get(f"{url_bill}&page={page}", timeout=10)
                data = res.json()
                break
            except requests.exceptions.ReadTimeout:
                print(f"  ⚠️ [page {page}] 타임아웃, 재시도 {retry + 1}/3")
                time.sleep(2)
        else:
            print(f"  ❌ [page {page}] 요청 실패, 건너뜀")
            continue

        for vote in data.get('rows', []):
            if vote.get("voteResultCode") != "찬성":
                continue

            member_id = vote.get("nationalAssemblyMemberId")
            bill_name = vote.get("bill", {}).get("name")

            if not member_id or not bill_name:
                continue

            existing = existing_data.get(member_id, "")
            if bill_name in existing:
                continue

            updated = existing + ("," if existing else "") + bill_name
            existing_data[member_id] = updated
            updates[member_id] = updated

    except Exception as e:
        print(f"  ❗ [page {page}] 예외 발생: {e}")
        continue

# 3. DB 일괄 업데이트
print(f"\n✅ DB에 {len(updates)}건 업데이트 중...")
for member_id, bill_list in updates.items():
    cursor.execute(
        "UPDATE politicians SET bill_approved = %s WHERE str_id = %s",
        (bill_list, member_id)
    )

conn.commit()

cursor.close()
conn.close()
print("완료")